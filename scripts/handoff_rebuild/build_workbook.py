"""Rebuild HANDOFF package v3 — clean doctor view (medical-only).

Doctor view (Sheet 1 'review'):
  Auto cols (grey, 10): pmc_id, pmc_url, local_xml, disease_name, omim_ids,
                        age, sex, has_family_history,
                        case_excerpt, hpo_phenotypes_extracted
  Doctor fills (orange, 5): check1_diagnosis_match, check2_hpo_accurate,
                            hpo_phenotypes_clean, review_decision, reviewer_notes

Rows: 198 (= 250 raw − 52 auto-Check3-fail). Auto-rejected rows are dropped.

External (not in HANDOFF folder): auto_checks_join.json — keeps Check 3/4 verdicts
keyed by pmc_id so we can join with doctor's returned xlsx later.

Check 4 (rarity) policy:
  - Explicit ">1/1000" Orphadata class → fail
  - Anything else (including missing prevalence) → pass
    Rationale: Orphanet's inclusion criterion IS rare disease; absence of a
    *validated* prevalence number means undocumented-but-rare, not non-rare.
"""
from __future__ import annotations
import csv
import gzip
import json
import re
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path

ROOT = Path("/Users/yutianzhao/Desktop/RDAgentBenchmark")
HOLDOUT_DIR = ROOT / "data/pmc_oa_holdout"
OLD_HANDOFF = HOLDOUT_DIR / "HANDOFF"
NEW_HANDOFF = HOLDOUT_DIR / "HANDOFF_v3"
XML_CACHE = HOLDOUT_DIR / "03_xml"
PREVALENCE_XML = ROOT / "data/orphadata/en_product9_prev.xml"
CHECK3_JSONL = HOLDOUT_DIR / "auto_check3_results.jsonl"
AUTO_JOIN_PATH = HOLDOUT_DIR / "auto_checks_join.json"

NEW_HANDOFF.mkdir(parents=True, exist_ok=True)
(NEW_HANDOFF / "pmc_fulltext").mkdir(exist_ok=True)


# ------------------------------------------------------------------
# Orphadata: explicit non-rare band only → fail Check 4
# ------------------------------------------------------------------
NON_RARE_BAND = ">1 / 1000"


def load_prevalence():
    print("Loading Orphadata prevalence XML…")
    tree = ET.parse(PREVALENCE_XML)
    root = tree.getroot()
    out = {}
    for d in root.iter("Disorder"):
        ocode = d.findtext("OrphaCode")
        name = d.findtext("Name") or ""
        dtype = d.findtext("DisorderType/Name") or ""
        bands = set()
        for prev in d.iter("Prevalence"):
            status = prev.findtext("PrevalenceValidationStatus/Name") or ""
            if status.strip().lower() != "validated":
                continue
            cls = prev.findtext("PrevalenceClass/Name") or ""
            if cls:
                bands.add(cls.strip())
        out[int(ocode)] = {
            "bands": list(bands),
            "disorder_type": dtype,
            "name": name,
        }
    print(f"  loaded {len(out)} disorders")
    return out


def check4_decision(prev_info: dict) -> tuple[str, str]:
    """Rare-disease auto-check.

    pass: Orphanet-registered AND no explicit non-rare band.
    fail: ONLY when Orphanet has explicit ">1/1000" entry.
    uncertain: Orphanet has no record at all (shouldn't happen — we matched).
    """
    if not prev_info:
        return "uncertain", "ORPHA code not found in Orphadata."
    bands = prev_info.get("bands", [])
    if NON_RARE_BAND in bands:
        return "fail", f"Orphadata classifies as NOT rare in Europe ({NON_RARE_BAND})."
    if bands:
        return "pass", f"Orphadata prevalence band(s): {', '.join(bands)}."
    return "pass", (
        f"Orphanet-registered rare disease (disorder_type='{prev_info.get('disorder_type', '')}'); "
        f"no validated prevalence — defaults to rare per Orphanet inclusion criterion."
    )


# ------------------------------------------------------------------
# PMC XML parser
# ------------------------------------------------------------------
PUB_DATE_RE = re.compile(
    r'<pub-date[^>]*(?:pub-type|date-type)="([^"]+)"[^>]*>(.*?)</pub-date>',
    re.DOTALL,
)


def _extract_pub_date(txt: str) -> str:
    matches = list(PUB_DATE_RE.findall(txt))
    priority = ["epub", "pub", "epub-ppub", "electronic", "pmc-release", "ppub", "collection"]
    def rank(t):
        try: return priority.index(t)
        except ValueError: return 99
    matches.sort(key=lambda x: rank(x[0]))
    for _, blk in matches:
        y = re.search(r"<year>(\d{4})</year>", blk)
        if not y:
            continue
        mo = re.search(r"<month>(\d{1,2})</month>", blk)
        d = re.search(r"<day>(\d{1,2})</day>", blk)
        out = y.group(1)
        if mo:
            out += f"-{int(mo.group(1)):02d}"
            if d:
                out += f"-{int(d.group(1)):02d}"
        return out
    m = re.search(r"<pub-date[^>]*>.*?<year>(\d{4})</year>.*?</pub-date>", txt, re.DOTALL)
    return m.group(1) if m else ""


FINAL_DX_PATTERNS = [
    r"\bfinal diagnosis\b",
    r"\bdefinitive diagnosis\b",
    r"\bdiagnosis was (?:established|confirmed|made)\b",
    r"\bwas (?:diagnosed|confirmed) (?:with|as)\b",
    r"\bconfirmed the diagnosis of\b",
    r"\bdiagnosis of [A-Z]",
]


def parse_pmc_xml(pmc_id: str) -> dict:
    path = XML_CACHE / f"PMC{pmc_id}.xml.gz"
    if not path.exists():
        return {"pub_pmc_date": "", "excerpt": "", "has_final_dx_string": False}
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as f:
        txt = f.read()
    pub_date = _extract_pub_date(txt)

    body_match = re.search(r"<body[^>]*>(.*?)</body>", txt, re.DOTALL)
    body_xml = body_match.group(1) if body_match else txt
    body_text = re.sub(r"<[^>]+>", " ", body_xml)
    body_text = re.sub(r"\s+", " ", body_text).strip()

    has_final = False
    excerpt = ""
    for pat in FINAL_DX_PATTERNS:
        m = re.search(pat, body_text, re.IGNORECASE)
        if m:
            has_final = True
            start = max(0, m.start() - 700)
            end = min(len(body_text), m.end() + 700)
            excerpt = body_text[start:end].strip()
            if start > 0: excerpt = "…" + excerpt
            if end < len(body_text): excerpt = excerpt + "…"
            break
    if not excerpt:
        excerpt = body_text[:1800].strip()
        if len(body_text) > 1800:
            excerpt += "…"
    if len(excerpt) > 2000:
        excerpt = excerpt[:1997] + "…"
    return {"pub_pmc_date": pub_date, "excerpt": excerpt, "has_final_dx_string": has_final}


# ------------------------------------------------------------------
def orpha_int(orpha_id: str):
    m = re.match(r"ORPHA:(\d+)", orpha_id or "")
    return int(m.group(1)) if m else None


def load_check3() -> dict:
    out = {}
    with open(CHECK3_JSONL) as f:
        for line in f:
            try:
                r = json.loads(line)
                out[r["pmc_id"]] = r
            except Exception:
                pass
    return out


def build_rows():
    print("Loading pool & old CSV…")
    pool = {}
    with open(OLD_HANDOFF / "candidates_full_pool.jsonl") as f:
        for line in f:
            rec = json.loads(line)
            pool[rec["pmc_id"]] = rec
    with open(OLD_HANDOFF / "review_template.csv") as f:
        csv_rows = list(csv.DictReader(f))

    prevalence = load_prevalence()
    check3 = load_check3()

    doctor_rows = []     # Check 3 pass — go in xlsx
    auto_join = {}       # pmc_id -> {auto_check3, auto_check4, ...} for later join

    print("Parsing 250 PMC XMLs…")
    for i, r in enumerate(csv_rows):
        pmc_id = r["pmc_id"]
        rec = pool.get(pmc_id, {})
        oint = orpha_int(rec.get("orpha_id", ""))
        prev_info = prevalence.get(oint, {}) if oint else {}
        pmc_info = parse_pmc_xml(pmc_id)
        c3 = check3.get(pmc_id, {})
        c4_dec, c4_reason = check4_decision(prev_info)

        auto_join[pmc_id] = {
            "orpha_id": rec.get("orpha_id", ""),
            "disease_name": rec.get("matched_orpha_name", ""),
            "pub_pmc_date": pmc_info["pub_pmc_date"],
            "prevalence_bands": prev_info.get("bands", []),
            "auto_check3_decision": c3.get("decision", "uncertain"),
            "auto_check3_reason": c3.get("reason", ""),
            "auto_check4_decision": c4_dec,
            "auto_check4_reason": c4_reason,
        }

        # Drop Check 3 fail from doctor view entirely
        if c3.get("decision") == "fail":
            continue

        doctor_rows.append({
            "pmc_id": pmc_id,
            "pmc_url": rec.get("pmc_url", f"https://pmc.ncbi.nlm.nih.gov/articles/PMC{pmc_id}/"),
            "local_xml": f"pmc_fulltext/PMC{pmc_id}.xml.gz",
            "disease_name": rec.get("matched_orpha_name", ""),
            "omim_ids": "; ".join(rec.get("omim_ids") or []),
            "age": rec.get("age_at_presentation_years"),
            "sex": rec.get("sex", ""),
            "has_family_history": rec.get("has_family_history", ""),
            "case_excerpt": pmc_info["excerpt"],
            "has_final_dx_string": "yes" if pmc_info["has_final_dx_string"] else "no",
            "hpo_phenotypes_extracted": "; ".join(rec.get("hpo_phenotypes") or []),
            # doctor-fill (raw labels only)
            "correct_diagnosis_if_wrong": "",
            "wrong_hpo_terms": "",
        })

    print(f"  doctor rows (Check3 pass): {len(doctor_rows)}")
    print(f"  auto_join keys: {len(auto_join)}")
    return doctor_rows, auto_join


# ------------------------------------------------------------------
def write_xlsx(rows):
    print("Writing xlsx…")
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    auto_cols = [
        "pmc_id", "pmc_url", "local_xml",
        "disease_name", "omim_ids",
        "age", "sex", "has_family_history",
        "case_excerpt", "has_final_dx_string",
        "hpo_phenotypes_extracted",
    ]
    doctor_cols = [
        "correct_diagnosis_if_wrong",
        "wrong_hpo_terms",
    ]
    # Friendly header labels with inline instruction
    header_labels = {
        "correct_diagnosis_if_wrong":
            "正确诊断 (LLM 抽错时填; 对则空)",
        "wrong_hpo_terms":
            "错误 HPO terms (用 ; 分隔; 全对则空)",
    }
    headers = auto_cols + doctor_cols
    display_headers = [header_labels.get(h, h) for h in headers]

    ws = wb.active
    ws.title = "review"
    ws.append(display_headers)

    auto_fill = PatternFill("solid", fgColor="305496")
    doctor_fill = PatternFill("solid", fgColor="C65911")
    autobody = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill = doctor_fill if h in doctor_cols else auto_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    n_rows = len(rows)
    excerpt_idx = headers.index("case_excerpt") + 1
    hpo_idx = headers.index("hpo_phenotypes_extracted") + 1

    for col_idx, h in enumerate(headers, 1):
        is_auto = h in auto_cols
        for row_idx in range(2, n_rows + 2):
            c = ws.cell(row=row_idx, column=col_idx)
            if is_auto:
                c.fill = autobody
            if col_idx in (excerpt_idx, hpo_idx):
                c.alignment = wrap

    widths = {
        "pmc_id": 11, "pmc_url": 22, "local_xml": 26,
        "disease_name": 28, "omim_ids": 18,
        "age": 6, "sex": 7, "has_family_history": 11,
        "case_excerpt": 65, "has_final_dx_string": 12,
        "hpo_phenotypes_extracted": 45,
        "correct_diagnosis_if_wrong": 32, "wrong_hpo_terms": 40,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    for row_idx in range(2, n_rows + 2):
        ws.row_dimensions[row_idx].height = 100
    # taller header row so the inline instructions wrap nicely
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "B2"
    # also wrap doctor columns
    for col_idx, h in enumerate(headers, 1):
        if h in doctor_cols:
            for row_idx in range(2, n_rows + 2):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap

    # Sheet 2: demo_examples (raw-label schema demo)
    ws2 = wb.create_sheet("demo_examples")
    dh = ["pmc_id", "disease_name (LLM 抽)", "正确诊断 (空=对)", "错误 HPO terms (空=全对)", "备注"]
    ws2.append(dh)
    demo_records = [
        ["13074162", "Werner syndrome", "", "", "LLM 抽得对; HPO 全是病人本身的"],
        ["10766305", "Lhermitte-Duclos disease", "", "", "诊断对 + HPO 全对"],
        ["10767403", "3M syndrome", "", "high-pitched voice; bird-like facies",
         "LLM 把别处描述的 phenotype 误入 list"],
        ["10768362", "Mevalonate kinase deficiency",
         "(reject — 这是 2018 年管理的 retrospective case series, 不是新报告)",
         "",
         "Check 1 没问题但本应被 cutoff filter 拒(已自动)"],
        ["EXAMPLE-DX-WRONG", "AMN (= Adrenomyeloneuropathy ABCD1)",
         "Acute macular neuroretinopathy", "",
         "同名异病: LLM 误映射到错的 ORPHA"],
        ["EXAMPLE-PARTIAL-HPO", "Marfan syndrome", "Loeys-Dietz syndrome",
         "ectopia lentis; aortic root dilatation",
         "诊断改了 + 2 个 HPO 实际属于鉴别诊断"],
    ]
    for r in demo_records:
        ws2.append(r)
    for i in range(1, len(dh) + 1):
        c = ws2.cell(row=1, column=i)
        c.fill = auto_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 38
    for row_idx in range(2, len(demo_records) + 2):
        ws2.row_dimensions[row_idx].height = 50
        for ci in (3, 4, 5):
            ws2.cell(row=row_idx, column=ci).alignment = wrap
    ws2.row_dimensions[1].height = 40
    ws2.freeze_panes = "A2"

    out_path = NEW_HANDOFF / "review_workbook.xlsx"
    wb.save(out_path)
    print(f"  wrote {out_path}")


def copy_xmls(rows):
    print("Copying PMC XML for doctor rows…")
    dst = NEW_HANDOFF / "pmc_fulltext"
    n = 0
    for r in rows:
        src = XML_CACHE / f"PMC{r['pmc_id']}.xml.gz"
        if src.exists():
            shutil.copy(src, dst / src.name)
            n += 1
    print(f"  copied {n} XML files")


def save_auto_join(auto_join):
    AUTO_JOIN_PATH.write_text(json.dumps(auto_join, indent=2))
    print(f"  saved auto join data → {AUTO_JOIN_PATH} ({len(auto_join)} entries)")


def main():
    rows, auto_join = build_rows()
    write_xlsx(rows)
    copy_xmls(rows)
    save_auto_join(auto_join)

    # diagnostics
    from collections import Counter
    c4_distrib = Counter()
    for pmc_id, info in auto_join.items():
        if info["auto_check3_decision"] != "pass":
            continue  # row dropped from doctor view
        c4_distrib[info["auto_check4_decision"]] += 1
    print("\n=== Final HANDOFF_v3 ===")
    print(f"  doctor xlsx rows:           {len(rows)}")
    print(f"  Check 3 pass + Check 4 pass: {c4_distrib.get('pass', 0)}")
    print(f"  Check 3 pass + Check 4 fail: {c4_distrib.get('fail', 0)}  (will auto-reject at join)")
    print(f"  Check 3 pass + Check 4 unc:  {c4_distrib.get('uncertain', 0)}")
    print(f"\n  Auto-reject (Check 3 fail) hidden from doctor: {len(auto_join) - len(rows)}")


if __name__ == "__main__":
    main()
